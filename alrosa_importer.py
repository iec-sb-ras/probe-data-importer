import base64
import hashlib
import os
import os.path
import re
import time
import uuid
from pprint import pprint

import openpyxl
import pandas as pd
import pudb
from openpyxl.cell import Cell
from pandas.core.internals.managers import new_block
from rdflib import (
    DCTERMS,
    FOAF,
    RDF,
    RDFS,
    XSD,
    BNode,
    Graph,
    Literal,
    URIRef,
)
from rdflib.namespace import SDO, WGS

from alrosa_convert_features import canonicalize_keys, convert_features_to_rdf
from namespace import BIBO, CGI, DBP, DBP_OWL, GS, MT, PT, SCHEMA, P

# Создание графа
G = Graph()

# Добавление алиасов (префиксов) для пространств имен
G.bind("foaf", FOAF)
G.bind("xsd", XSD)
G.bind("rdf", RDF)
G.bind("rdfs", RDFS)
G.bind("dcterms", DCTERMS)
G.bind("wgs", WGS)
G.bind("sdo", SDO)
G.bind("pt", PT)
G.bind("p", P)
G.bind("schema", SCHEMA)
G.bind("bibo", BIBO)
G.bind("mt", MT)
G.bind("gs", GS)
G.bind("cgi", CGI)
G.bind("dbp", DBP)
G.bind("dbp-owl", DBP_OWL)

import pickle


def save_dict_as_pickle(data_dict, filename="data.pkl"):
    """
    Сохраняет весь словарь целиком в pickle файл
    """
    with open(filename, "wb") as f:
        pickle.dump(data_dict, f)
    print(f"Словарь сохранен в {filename}")


def load_dict_from_pickle(filename="data.pkl"):
    """
    Загружает словарь из pickle файла
    """
    with open(filename, "rb") as f:
        return pickle.load(f)


# Использование:
# save_dict_as_pickle(data_dict, "backup.pkl")

# Загрузить обратно:
# loaded_dict = load_dict_from_pickle("backup.pkl")


def search_substring_on_sheet(sheet, column_number, substring):
    """Search for substring in a specific column of an openpyxl sheet."""
    matches = []
    for row in sheet.iter_rows(min_col=column_number, max_col=column_number):
        cell = row[0]
        if cell.value and substring.lower() in str(cell.value).lower():
            matches.append((cell.row, cell.value))
    if not matches:
        print("ERROR: table {} is not found".format(substring))
    return matches


def clean_value_of_cell(cell):
    """Extract and clean cell value, returning appropriate Python type."""
    if isinstance(cell, Cell):
        value = cell.value
    else:
        value = cell

    if value is None:
        return None

    # Handle numeric types
    if isinstance(value, (int, float)):
        return value

    # Handle string values
    if isinstance(value, str):
        cleaned = value.strip()
        lc = cleaned.lower()

        if lc in ["н.д.", ""]:
            return None

        # Try to convert to int or float if possible
        try:
            # Check for integer
            if cleaned.isdigit():
                return int(cleaned)
            # Check for float
            cleaned = cleaned.replace(",", ".")
            return float(cleaned)
        except (ValueError, AttributeError):
            # Return cleaned string if conversion fails
            return cleaned

    # Return other types as-is (datetime, etc.)
    return value


def append_features(a_dict, name_row, value_row, skip_zeros=False):
    """
    Append cleaned feature values from two rows to a dictionary.

    Args:
        a_dict: словарь для заполнения
        name_row: строка с названиями параметров
        value_row: строка со значениями
        skip_zeros: если True, пропускает нулевые значения (0 и 0.0)
    """
    if name_row and value_row:
        for name_cell, value_cell in zip(name_row, value_row):
            if value_cell.value is None:
                continue
            key = str(name_cell.value).strip()
            val = clean_value_of_cell(value_cell)

            if val is not None:
                # Проверяем на ноль только если нужно пропускать
                if skip_zeros and isinstance(val, (int, float)) and val == 0:
                    continue
                a_dict[key] = val


def import_head_title_one_row_features(
    sheet, column_number, match_string, special=False
) -> dict:
    """Extract features from the row containing match_string in the first column."""
    features = {}

    matches = search_substring_on_sheet(sheet, column_number, match_string)
    if not matches:
        print("ERROR: {} not found in this {}.".format(match_string, sheet.title))
        return features  # Return empty dict if no match

    row_num, _ = matches[0]  # Use first match
    row = sheet[row_num]

    # Read next row for feature names
    if special:
        next_row = sheet[row_num] if row_num <= sheet.max_row else None
        # Read row after that for feature values
        value_row = sheet[row_num + 1] if row_num + 1 <= sheet.max_row else None
    else:
        next_row = sheet[row_num + 1] if row_num + 1 <= sheet.max_row else None
        # Read row after that for feature values
        value_row = sheet[row_num + 2] if row_num + 2 <= sheet.max_row else None

    append_features(features, next_row, value_row)

    return features


def import_head_title_multirow_as_data_frame(sheet, column_number, match_string):
    """Import multi-row data starting from the row after match_string as a pandas DataFrame."""
    matches = search_substring_on_sheet(sheet, column_number, match_string)
    if not matches:
        return None

    start_row, _ = matches[0]  # Use first match
    data_start_row = start_row + 1  # Data begins on the next row

    # Read data into a list of lists
    data = []
    for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
        if any(cell is not None for cell in row):  # Stop at first completely empty row
            data.append(row)
        else:
            break

    if not data:
        return None

    # Use first row as headers, rest as data
    df = pd.DataFrame(data[1:], columns=data[0])
    return df


def import_transposed_table_as_data_frame(sheet, column_number, match_string):
    """
    Import transposed table where:
    - Row with match_string contains headers (parameters)
    - Subsequent rows contain data for different samples
    - Columns represent different samples/analyses
    """
    # pu.db
    matches = search_substring_on_sheet(sheet, column_number, match_string)
    if not matches:
        return None

    header_row, _ = matches[0]  # Row with parameter names
    data_start_row = header_row + 1  # Data starts from next row

    # Read all data rows until empty
    data_rows = []
    for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
        if any(cell is not None for cell in row):
            data_rows.append(row)
        else:
            break

    if not data_rows:
        return None

    # Get headers (parameter names) from the first data column
    # First column contains parameter names, subsequent columns contain sample data
    headers = []

    for i, row in enumerate(data_rows):
        cell = row[0]
        if cell and str(cell).strip():
            headers.append(str(cell).strip())
        else:
            headers.append("val_{}".format(i + 1))

    df = pd.DataFrame()

    # print(headers)
    # print(data_rows[0])
    for i, row in enumerate(data_rows):
        dt = row[1:]
        dt = [clean_value_of_cell(c) for c in dt]
        df[headers[i]] = dt

    if not df.empty and df.size > 0:
        return df
    else:
        return None


# Smalltalk inheritance, be aswre ;-)


def remove_empty_columns_from_data_frame(df):
    """Remove columns from a DataFrame where all values are NaN/None."""
    return df.dropna(axis=1, how="all")


def add_to_dict_if_not_none(main_dict, name, an_object):
    """Add a dictionary to main_dict under key 'name' if a_dict is not empty."""
    if isinstance(an_object, pd.DataFrame):
        an_object = remove_empty_columns_from_data_frame(an_object)
    if an_object is not None:
        main_dict[name] = an_object
    else:
        print("ERROR: {} object is None, do not add it".format(name))


def import_excel_table_into_dict(workbook, sheet_name_or_index):
    """
    Import data from Excel file into a dictionary using openpyxl.

    Args:
        file_path: Path to the Excel file
        sheet_name_or_index: Sheet name (str) or index (int)

    Returns:
        dict: Dictionary with row numbers as keys and row data as dicts
    """

    # Get the sheet
    if isinstance(sheet_name_or_index, str):
        sheet = workbook[sheet_name_or_index]
    else:
        sheet = workbook.worksheets[sheet_name_or_index]

    # print sheet Name
    print(f"Processing sheet: {sheet.title}")

    matches = search_substring_on_sheet(sheet, 0, "НДС")
    if matches:
        print("This sheet is not a tube")
        return sheet, None

    data_dict = {}

    features = {}

    tube_aim_features = import_head_title_one_row_features(
        sheet, 1, "Целевой показатель"
    )
    add_to_dict_if_not_none(features, "target", tube_aim_features)

    geology_aim_features = import_head_title_one_row_features(sheet, 1, "Геология")
    add_to_dict_if_not_none(features, "geology", geology_aim_features)

    olivine_aim_features = import_head_title_one_row_features(
        sheet, 1, "Оливины I-генерации"
    )
    add_to_dict_if_not_none(features, "olivine", olivine_aim_features)

    aa_aim_features = import_head_title_one_row_features(
        sheet, 1, "алмазная ассоциация", special=True
    )
    add_to_dict_if_not_none(features, "assoc", aa_aim_features)

    add_to_dict_if_not_none(data_dict, "features", features)

    dfs = {}

    phlogopite = import_head_title_multirow_as_data_frame(
        sheet, 1, "Состав флогопита из основной массы"
    )
    add_to_dict_if_not_none(dfs, "phlogopite", phlogopite)

    isotopic_contents = import_head_title_multirow_as_data_frame(
        sheet, 1, "Изотопный состав"
    )
    add_to_dict_if_not_none(dfs, "isotopic", isotopic_contents)

    epma_contents = import_head_title_multirow_as_data_frame(
        sheet, 1, "EPMA составы минералов"
    )
    add_to_dict_if_not_none(dfs, "epma", epma_contents)

    lam_contents = import_head_title_multirow_as_data_frame(
        sheet, 1, "LAM ICP составы гранатов"
    )
    add_to_dict_if_not_none(dfs, "lam", lam_contents)

    micro_diamonds = import_head_title_multirow_as_data_frame(sheet, 1, "МИКРОАЛМАЗЫ")
    add_to_dict_if_not_none(dfs, "diamonds", micro_diamonds)

    micro_oxides = import_head_title_multirow_as_data_frame(sheet, 1, "МИКРООКСИДЫ")
    add_to_dict_if_not_none(dfs, "oxides", micro_oxides)

    petrochemistry = import_transposed_table_as_data_frame(sheet, 1, "Петрохимия")
    add_to_dict_if_not_none(dfs, "petrochemy", petrochemistry)

    geochemy = import_transposed_table_as_data_frame(sheet, 1, "Геохимия")
    add_to_dict_if_not_none(dfs, "geochemy", petrochemistry)

    add_to_dict_if_not_none(data_dict, "frames", dfs)

    return sheet, data_dict


def search_file_to_root(name):
    """
    Search for an Excel file by name in the current directory and its parent directories.

    Args:
        name (str): Name of the Excel file to search for.

    Returns:
        str or None: Full path to the Excel file if found, None otherwise.
    """
    current_dir = os.path.abspath(os.path.dirname(__file__))

    while True:
        # Check if file exists in current directory
        file_path = os.path.join(current_dir, name)
        if os.path.isfile(file_path):
            return file_path

        # Move to parent directory
        parent_dir = os.path.dirname(current_dir)
        if parent_dir == current_dir:  # Reached root directory
            break
        current_dir = parent_dir

    return None


keymaster = {}


def keymaster_update(section_key, data_section, with_pipe_names=False, pipe_name=None):

    if with_pipe_names:
        a_set = keymaster.setdefault(section_key, dict())
    else:
        a_set = keymaster.setdefault(section_key, set())

    if isinstance(data_section, list):
        update_list = data_section
    else:
        update_list = data_section.keys()

    if with_pipe_names:
        for item in update_list:
            s = a_set.setdefault(item, set())
            s.add(pipe_name)
    else:
        a_set.update(update_list)


def convert_to_canonic_form_features(features):
    new_features = {}
    new_features.update(features)
    new_features = canonicalize_keys(new_features)
    return new_features


def normalize_columns(df, table_name):
    """
    Нормализует имена столбцов DataFrame

    Parameters:
    df: pandas DataFrame
    table_name: str, название таблицы для логирования

    Returns:
    DataFrame с нормализованными именами столбцов
    """
    new_columns = []

    for i, col in enumerate(df.columns):
        # Пропускаем None
        if pd.isna(col) or col is None:
            new_col = f"val_{i}"
            print(f"  {table_name}: колонка {i} была None -> '{new_col}'")
            new_columns.append(new_col)
            continue

        # Преобразуем в строку и удаляем пробелы в начале/конце
        col_str = str(col).strip()
        original = col_str

        # Удаляем \n и лишние пробелы внутри
        col_str = " ".join(col_str.split())
        col_str = col_str.replace("\n", "")

        # Заменяем / на _
        col_str = col_str.replace("/", "_")

        # Заменяем # на num
        col_str = col_str.replace("#", "num")

        # Заменяем ε на eps
        col_str = col_str.replace("ε", "eps")

        # Заменяем остальные проблемные символы на _
        col_str = re.sub(r"[^\w\d\s]", "_", col_str)
        col_str = re.sub(r"\s+", "_", col_str)

        # Убираем множественные подчёркивания
        col_str = re.sub(r"_+", "_", col_str)

        # Убираем подчёркивания в начале и конце
        col_str = col_str.strip("_")

        # Если после всех преобразований получилась пустая строка
        if not col_str:
            col_str = f"val_{i}"
            print(f"  {table_name}: колонка '{original}' стала пустой -> '{col_str}'")

        # Проверка на дубликаты (если такое имя уже есть)
        base_col = col_str
        counter = 1
        while col_str in new_columns:
            col_str = f"{base_col}_{counter}"
            counter += 1

        if original != col_str:
            print(f"  {table_name}: '{original}' -> '{col_str}'")

        new_columns.append(col_str)

    df.columns = new_columns
    return df


def convert_to_canonic_form_frames(frames, pipe_name):
    """Convert all dataframes to canonical form."""
    new_frames = {}
    new_frames.update(frames)
    frame_names = [
        "phlogopite",
        "isotopic",
        "epma",
        "lam",
        "diamonds",
        "oxides",
        "petrochemy",
        "geochemy",
    ]
    for fn in frame_names:
        df = frames[fn]
        df = normalize_columns(df, pipe_name + ":" + fn)
        column_names = df.columns.tolist()
        keymaster_update(fn, column_names, pipe_name=pipe_name, with_pipe_names=True)
    return new_frames


def process_excel_file(file_path):
    """
    Process an Excel file and extract all tube data.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        dict: Dictionary with sheet names as keys and extracted data as values.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_data = {}

    for sheet in wb.worksheets:
        sheet_name = sheet.title
        print(f"Processing sheet: {sheet_name}")

        # Check if this sheet contains tube data
        matches = search_substring_on_sheet(sheet, 1, "НДС")
        if matches:
            print(f"  Skipping sheet '{sheet_name}' - not a tube sheet")
            continue

        # Import data from this sheet
        _, data_dict = import_excel_table_into_dict(wb, sheet_name)

        if data_dict:
            # Convert to canonical form
            canonical_data = convert_to_canonic_form(data_dict)
            all_data[sheet_name] = canonical_data
            print(f"  Successfully extracted data from '{sheet_name}'")
        else:
            print(f"  No data extracted from '{sheet_name}'")

    return all_data


def main():
    """Main function to search for and process Excel files."""
    # Search for Excel files
    excel_files = [
        "tubes.xlsx",
        "data.xlsx",
        "kimberlite_tubes.xlsx",
        "kimberlite_data.xlsx",
    ]

    found_file = None
    for file_name in excel_files:
        file_path = search_file_to_root(file_name)
        if file_path:
            found_file = file_path
            print(f"Found Excel file: {found_file}")
            break

    if not found_file:
        print("No Excel files found in current or parent directories.")
        return

    # Process the Excel file
    print(f"Processing file: {found_file}")
    all_data = process_excel_file(found_file)

    # Save the results
    if all_data:
        output_file = "extracted_tube_data.pkl"
        save_dict_as_pickle(all_data, output_file)
        print(f"Data saved to {output_file}")

        # Print summary
        print(f"\nSummary:")
        print(f"  Total sheets processed: {len(all_data)}")
        for sheet_name, data in all_data.items():
            features_count = len(data.get("features", {}))
            frames_count = len(data.get("frames", {}))
            print(
                f"  {sheet_name}: {features_count} features, {frames_count} dataframes"
            )
    else:
        print("No tube data found in the Excel file.")


def convert_to_canonic_form(tube):
    """
    Convrts tube data in form of dictionary and pandas dataframe in the same structurebut ranamed keys into canonic format
    """

    name, data = tube
    if data is None:
        return name, data

    features = data.get("features", {})
    features = convert_to_canonic_form_features(features)
    frames = data.get("frames", {})
    frames = convert_to_canonic_form_frames(frames, pipe_name=name)

    new_data = {"features": features, "frames": frames}
    return name, new_data


def generate_deterministic_uuid(
    pipe_name: str = "",
    namespace: str = "http://crust.irk.ru/ontology/contents/1.0/",
    pipe_uuid: str = "",
    **kwargs,
) -> uuid.UUID:
    """
    Генерирует детерминированный UUID v5 на основе имени трубки (и/или UUID) и дополнительных данных
    """
    # Создаем UUID v5 на основе пространства имен и имени
    namespace_uuid = uuid.NAMESPACE_DNS  # Или можно создать свой
    uuid_additional_data = ""
    for key, value in kwargs.items():
        uuid_additional_data += f"{str(key)}:{str(value)}"
    new_pipe_uuid = uuid.uuid5(
        namespace_uuid, f"{namespace}{pipe_name}{str(pipe_uuid)}{uuid_additional_data}"
    )
    return new_pipe_uuid


def export_tube(g, tube):
    tube_name, tube_dict = tube

    print("Processing pipe {}".format(tube_name))

    if tube_dict is None:  # ценник, really
        return None

    tube_uri = P[tube_name]

    # Add type assertion
    g.add((tube_uri, RDF.type, PT.KimberlitePipe))

    # Генерируем детерминированный UUID
    pipe_uuid = generate_deterministic_uuid(tube_name)

    tube_dict["UUID"] = str(pipe_uuid)

    g.add((tube_uri, PT.uuid, Literal(pipe_uuid, datatype=XSD.string)))

    features = tube_dict.get("features", {})

    convert_features_to_rdf(G, (tube_name, features), tube_uri)

    print("INFO: features after conversion:", end=": ")
    pprint(features)

    return tube_uri


def main():
    # Example usage

    file_path = "data/tubes.xlsx"
    tubes_path = "tubes.pkl"

    tubes_pn = search_file_to_root(tubes_path)
    if tubes_pn is not None:
        start_time = time.time()
        tubes = load_dict_from_pickle(tubes_pn)
        load_time = time.time() - start_time
        print(f"INFO: Load success! Time: {load_time:.2f} sec")
        del tubes["ценник"]
    else:
        print("INFO: Starting import from excel")
        tubes = {}
        workbook = openpyxl.load_workbook(
            search_file_to_root(file_path), data_only=True
        )

        # for sheet_number in [1]:
        for sheet_number in range(len(workbook.sheetnames)):
            sheet, excel_data = import_excel_table_into_dict(workbook, sheet_number)
            # pprint(excel_data)
            tubes[sheet.title.strip()] = excel_data

        save_dict_as_pickle(tubes, tubes_path)
        print("INFO: Conversionhas been done. Rerun if export needed.")
        quit()

    for tube_item in tubes.items():
        tube_item = convert_to_canonic_form(tube_item)
        export_tube(G, tube_item)
        # break

    pprint(keymaster)

    # save Graph G in ../gql-server/fuseki/a-box.ttl

    output_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        "gql-server",
        "fuseki",
        "a-box.ttl",
    )
    # G.serialize(destination=output_path, format="turtle")
    # print(f"RDF graph saved to {output_path}")

    output_path = "a-box.ttl"
    # G.serialize(destination=output_path, format="turtle")
    # print(f"RDF graph saved to {output_path}")


if __name__ == "__main__":
    main()
